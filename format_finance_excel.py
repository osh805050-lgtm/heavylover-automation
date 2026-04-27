# -*- coding: utf-8 -*-
"""
헤비로버 재정조사 엑셀 파일 가독성 개선 스크립트
- 수식/계산식은 절대 변경하지 않음 (스타일만 적용)
- 색상으로 섹션 구분, 테두리, 정렬, 글꼴 크기 조정
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from copy import copy

SRC = r"C:\Users\osh80\OneDrive\바탕 화면\사업\운영예산\재정조사\26.4.22 재정조사 - 복사본.xlsx"

# ========== 색상 팔레트 ==========
NAVY        = "1F4E78"   # 메인 헤더 (진한 남색)
BLUE        = "2E75B6"   # 섹션 제목
LIGHT_BLUE  = "DDEBF7"   # 헤더 보조
PALE_BLUE   = "F2F8FC"   # 줄무늬 배경
YELLOW      = "FFF2CC"   # 합계/총계
LIGHT_YEL   = "FFF9E6"   # 부분합
GREEN       = "E2EFDA"   # 목표/중요 수치
LIGHT_GRAY  = "F2F2F2"   # 비고/보조 정보
ORANGE      = "FCE4D6"   # 강조 경고
WHITE       = "FFFFFF"

# ========== 공통 스타일 요소 ==========
thin_gray = Side(border_style="thin", color="BFBFBF")
medium_navy = Side(border_style="medium", color=NAVY)

border_all = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
border_bottom_thick = Border(
    left=thin_gray, right=thin_gray, top=thin_gray, bottom=medium_navy
)

def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def font_header(size=11, color=WHITE, bold=True):
    return Font(name="맑은 고딕", size=size, color=color, bold=bold)

def font_body(size=10, color="000000", bold=False, italic=False):
    return Font(name="맑은 고딕", size=size, color=color, bold=bold, italic=italic)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_cell(cell, *, fill_color=None, font=None, alignment=None, border=None,
               number_format=None):
    """셀의 값(수식 포함)은 건드리지 않고 스타일만 적용."""
    if fill_color is not None:
        cell.fill = fill(fill_color)
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def style_range(ws, cell_range, **kwargs):
    for row in ws[cell_range]:
        for cell in row:
            style_cell(cell, **kwargs)


def apply_header(ws, cell_range):
    """메인 헤더 스타일 (진한 남색 + 흰색 굵은 글씨)."""
    style_range(
        ws, cell_range,
        fill_color=NAVY, font=font_header(size=11),
        alignment=center, border=border_all,
    )


def apply_subheader(ws, cell_range):
    """서브헤더 (밝은 파랑 + 굵은 글씨)."""
    style_range(
        ws, cell_range,
        fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
        alignment=center, border=border_all,
    )


def apply_data(ws, cell_range, *, currency=True, highlight=None):
    """일반 데이터 영역."""
    fmt = '#,##0"원"' if currency else None
    for row in ws[cell_range]:
        for cell in row:
            style_cell(
                cell,
                fill_color=highlight,
                font=font_body(size=10),
                alignment=right_wrap,
                border=border_all,
                number_format=fmt,
            )


def set_row_heights(ws, heights):
    for row, h in heights.items():
        ws.row_dimensions[row].height = h


def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ========== 파일 로드 ==========
wb = openpyxl.load_workbook(SRC)
sheets = wb.sheetnames
print("시트 목록:", sheets)

# 시트 인덱스로 접근 (이름이 깨져있어도 순서로 처리)
# 순서: 1월 재정조사 현황, 일별 입금금액, 1월 돈흐름,
#      월별 재구매타겟, 월별 신규구매타겟, 월별 목표달성, 월별 매출 데이터

# ========== 시트 1: 월 재정조사 현황 ==========
ws = wb[sheets[0]]
# 전체 기본 글꼴 초기화
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.font is None or not cell.font.name:
            cell.font = font_body()

# 상단 자산/미납 테이블 (A1:H14)
# 헤더 (1행)
for coord in ["A1", "B1", "C1", "F1", "G1", "H1"]:
    style_cell(ws[coord], fill_color=NAVY, font=font_header(size=11),
               alignment=center, border=border_all)

# 항목/금액 영역 (A2:C14) - 자산
for row_num in range(2, 15):
    # 항목
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    # 금액
    style_cell(ws.cell(row=row_num, column=2),
               fill_color=PALE_BLUE if row_num % 2 == 0 else WHITE,
               font=font_body(size=10),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    # 비고
    style_cell(ws.cell(row=row_num, column=3),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 핵심 합계 행 강조 (B12=전체 현금, B13=전체 자산, B14=전체-미납)
for row_num in [12, 13, 14]:
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=YELLOW, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    style_cell(ws.cell(row=row_num, column=2),
               fill_color=YELLOW, font=font_body(size=11, bold=True, color="C00000"),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')

# 미납금 영역 (F2:H9)
for row_num in range(2, 10):
    style_cell(ws.cell(row=row_num, column=6),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    style_cell(ws.cell(row=row_num, column=7),
               fill_color=PALE_BLUE if row_num % 2 == 0 else WHITE,
               font=font_body(size=10),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    style_cell(ws.cell(row=row_num, column=8),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 미납 합계 행 (G9)
style_cell(ws.cell(row=9, column=6),
           fill_color=ORANGE, font=font_body(size=10, bold=True),
           alignment=left_wrap, border=border_all)
style_cell(ws.cell(row=9, column=7),
           fill_color=ORANGE, font=font_body(size=11, bold=True, color="C00000"),
           alignment=right_wrap, border=border_all,
           number_format='#,##0"원"')

# 단가 테이블 헤더 (A16:I16)
apply_header(ws, "A16:I16")

# 단가 데이터 (A17:I21)
for row_num in range(17, 22):
    for col in range(1, 10):
        bg = PALE_BLUE if row_num % 2 == 0 else WHITE
        fnt = font_body(size=10)
        al = right_wrap if col > 1 else left_wrap
        fmt = '#,##0' if col > 1 else None
        style_cell(ws.cell(row=row_num, column=col),
                   fill_color=bg, font=fnt, alignment=al,
                   border=border_all, number_format=fmt)

# 설명 영역 (A23:A26) - 노트
for row_num in range(23, 27):
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

style_cell(ws["B26"], fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
           alignment=left_wrap, border=border_all)

# 세무영 관련 (A28) + 두번째 단가 테이블
style_cell(ws["A28"], fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
           alignment=left_wrap, border=border_all)

# 할당량 테이블 (O27:U43) - 오른쪽 영역
apply_subheader(ws, "O27:U27")
for row_num in range(28, 33):
    for col_letter in ["O", "P", "Q", "R", "S", "T", "U"]:
        cell = ws[f"{col_letter}{row_num}"]
        bg = PALE_BLUE if row_num % 2 == 0 else WHITE
        style_cell(cell, fill_color=bg, font=font_body(size=10),
                   alignment=center, border=border_all)

# 두 번째 단가 테이블 헤더 (A33:I33)
apply_header(ws, "A33:I33")

# 두 번째 단가 데이터 (A34:I40)
for row_num in range(34, 41):
    for col in range(1, 10):
        bg = PALE_BLUE if row_num % 2 == 0 else WHITE
        al = right_wrap if col > 1 else left_wrap
        fmt = '#,##0' if col > 1 else None
        style_cell(ws.cell(row=row_num, column=col),
                   fill_color=bg, font=font_body(size=10),
                   alignment=al, border=border_all, number_format=fmt)

# 오른쪽 설명 표 (S36:T43)
for row_num in range(36, 44):
    for col_letter in ["O", "P", "Q", "R", "S", "T", "U", "V", "Y", "Z"]:
        cell = ws[f"{col_letter}{row_num}"]
        if cell.value is not None:
            style_cell(cell, fill_color=LIGHT_GRAY, font=font_body(size=9),
                       alignment=center, border=border_all)

# 컬럼 너비
set_col_widths(ws, {
    "A": 28, "B": 15, "C": 45, "D": 14, "E": 14,
    "F": 14, "G": 16, "H": 40, "I": 16,
    "O": 10, "P": 10, "Q": 10, "R": 10, "S": 10, "T": 10, "U": 12,
    "V": 12, "Y": 10, "Z": 12,
})

# 행 높이 - 줄바꿈 있는 긴 텍스트 셀은 충분히 높게
set_row_heights(ws, {
    1: 28,
    2: 60,   # 토스 페이먼츠 (\n 포함)
    3: 60,   # 스마트스토어 (\n 포함)
    4: 90,   # 쿠팡 비고 (5줄 이상)
    5: 32,
    6: 32,
    7: 60,   # 카카오페이 비고 (\n 포함)
    8: 28,
    9: 32,
    10: 28,
    11: 28,
    12: 28,
    13: 28,
    14: 32,
    16: 28,
    33: 28,
})

ws.freeze_panes = "A2"

# ========== 시트 2: 일별 입금금액 ==========
ws = wb[sheets[1]]

# 헤더 (A1:G1)
apply_header(ws, "A1:G1")

# 일별 데이터 (A2:G21)
for row_num in range(2, 22):
    # 날짜
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=center, border=border_all)
    # 플랫폼 금액 (B~E)
    for col in range(2, 6):
        bg = PALE_BLUE if row_num % 2 == 0 else WHITE
        style_cell(ws.cell(row=row_num, column=col),
                   fill_color=bg, font=font_body(size=10),
                   alignment=right_wrap, border=border_all,
                   number_format='#,##0"원"')
    # 합계 (F)
    style_cell(ws.cell(row=row_num, column=6),
               fill_color=LIGHT_YEL, font=font_body(size=10, bold=True),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    # 비고 (G)
    style_cell(ws.cell(row=row_num, column=7),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 월 합계 행 (A22:G22)
for col in range(1, 8):
    fmt = '#,##0"원"' if col > 1 and col < 7 else None
    if col == 6:
        fmt = '#,##0"원"'
    style_cell(ws.cell(row=22, column=col),
               fill_color=YELLOW, font=font_body(size=11, bold=True, color="C00000"),
               alignment=center if col == 1 else right_wrap,
               border=border_all, number_format=fmt)

# 두번째 테이블 헤더 (A24:G24)
apply_header(ws, "A24:G24")

# 두번째 테이블 데이터 (A25:G27)
for row_num in range(25, 28):
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    for col in range(2, 8):
        bg = PALE_BLUE if row_num % 2 == 0 else WHITE
        fmt = '#,##0"원"' if col in [3, 4, 5, 6, 7] else '#,##0'
        style_cell(ws.cell(row=row_num, column=col),
                   fill_color=bg, font=font_body(size=10),
                   alignment=right_wrap, border=border_all,
                   number_format=fmt)

# 세번째 작은 테이블 (A30:E32)
for row_num in range(30, 33):
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None or (row_num == 31 and col <= 5):
            if row_num == 30:
                style_cell(cell, fill_color=LIGHT_BLUE,
                           font=font_body(size=10, bold=True),
                           alignment=center, border=border_all)
            else:
                style_cell(cell, fill_color=WHITE,
                           font=font_body(size=10),
                           alignment=left_wrap if col == 1 else right_wrap,
                           border=border_all,
                           number_format='#,##0"원"' if col >= 3 else None)

# 요약 테이블 (A34:C41)
apply_header(ws, "A34:C34")
for row_num in range(35, 42):
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    # B: 금액
    bg = YELLOW if row_num in [38, 40] else (PALE_BLUE if row_num % 2 == 0 else WHITE)
    bold = row_num in [38, 40, 41]
    color = "C00000" if row_num in [38, 40] else "000000"
    style_cell(ws.cell(row=row_num, column=2),
               fill_color=bg,
               font=font_body(size=11 if bold else 10, bold=bold, color=color),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    # C: 비고
    style_cell(ws.cell(row=row_num, column=3),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 메모 영역 (A43:A46)
for row_num in range(43, 47):
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_GRAY if row_num > 43 else ORANGE,
               font=font_body(size=9, italic=True, bold=(row_num == 43)),
               alignment=left_wrap, border=border_all)

set_col_widths(ws, {
    "A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 20, "G": 30,
})
set_row_heights(ws, {1: 28, 24: 28, 34: 28, 35: 28, 36: 28, 37: 28,
                     38: 28, 39: 28, 40: 32, 41: 28})
ws.freeze_panes = "A2"

# ========== 시트 3: 1월 돈흐름 ==========
ws = wb[sheets[2]]

# 헤더 (A1, C1, F1, H1)
for coord in ["A1", "B1", "C1", "F1", "G1", "H1"]:
    style_cell(ws[coord], fill_color=NAVY, font=font_header(size=11),
               alignment=center, border=border_all)

# 왼쪽 자산 영역 (A2:C8)
for row_num in range(2, 9):
    style_cell(ws.cell(row=row_num, column=1),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    style_cell(ws.cell(row=row_num, column=2),
               fill_color=PALE_BLUE if row_num % 2 == 0 else WHITE,
               font=font_body(size=10),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    style_cell(ws.cell(row=row_num, column=3),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 전체현금 합계 (B8)
style_cell(ws["A8"], fill_color=YELLOW, font=font_body(size=10, bold=True),
           alignment=left_wrap, border=border_all)
style_cell(ws["B8"], fill_color=YELLOW,
           font=font_body(size=11, bold=True, color="C00000"),
           alignment=right_wrap, border=border_all,
           number_format='#,##0"원"')

# 오른쪽 미납 영역 (F2:H9)
for row_num in range(2, 10):
    style_cell(ws.cell(row=row_num, column=6),
               fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
               alignment=left_wrap, border=border_all)
    style_cell(ws.cell(row=row_num, column=7),
               fill_color=PALE_BLUE if row_num % 2 == 0 else WHITE,
               font=font_body(size=10),
               alignment=right_wrap, border=border_all,
               number_format='#,##0"원"')
    style_cell(ws.cell(row=row_num, column=8),
               fill_color=LIGHT_GRAY, font=font_body(size=9, italic=True),
               alignment=left_wrap, border=border_all)

# 미납 합계 (G9)
style_cell(ws["F9"], fill_color=ORANGE, font=font_body(size=10, bold=True),
           alignment=left_wrap, border=border_all)
style_cell(ws["G9"], fill_color=ORANGE,
           font=font_body(size=11, bold=True, color="C00000"),
           alignment=right_wrap, border=border_all,
           number_format='#,##0"원"')

set_col_widths(ws, {
    "A": 32, "B": 16, "C": 45, "D": 12, "E": 12,
    "F": 14, "G": 18, "H": 40,
})
set_row_heights(ws, {
    1: 28, 2: 48, 3: 60, 4: 32, 5: 32,
    6: 32, 7: 32, 8: 32, 9: 32,
})
ws.freeze_panes = "A2"

# ========== 시트 4: 월별 재구매타겟 ==========
ws = wb[sheets[3]]

# 1. 월 목표 섹션 (A2:H8)
# I1
style_cell(ws["I1"], fill_color=BLUE, font=font_header(size=10),
           alignment=center, border=border_all)

# B2:I3 상단 목표 설명
for row_num in range(2, 4):
    for col in range(2, 10):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            style_cell(cell, fill_color=LIGHT_BLUE,
                       font=font_body(size=10, bold=True),
                       alignment=center, border=border_all)

# 월 목표 매출 블록 (A3:E8)
for row_num in range(3, 9):
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            if col == 1:  # 월 이름
                style_cell(cell, fill_color=NAVY, font=font_header(size=11),
                           alignment=center, border=border_all)
            elif col == 2:  # 매출 목표
                style_cell(cell, fill_color=GREEN,
                           font=font_body(size=11, bold=True, color="375623"),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')
            elif col == 3:  # 설명
                style_cell(cell, fill_color=LIGHT_GRAY,
                           font=font_body(size=9, italic=True),
                           alignment=left_wrap, border=border_all)
            elif col == 4:  # 고객단가
                style_cell(cell, fill_color=PALE_BLUE, font=font_body(size=10),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')

# G2:I3 우측 목표 블록
for coord in ["G2", "H2", "I2", "G3", "H3", "I3", "F3", "G4", "F4"]:
    cell = ws[coord]
    if cell.value is not None:
        if coord in ["G2", "H2"]:
            style_cell(cell, fill_color=BLUE, font=font_header(size=10),
                       alignment=center, border=border_all)
        elif coord == "I2":
            style_cell(cell, fill_color=BLUE, font=font_header(size=10),
                       alignment=center, border=border_all)
        elif coord in ["G3", "H3"]:
            style_cell(cell, fill_color=LIGHT_BLUE,
                       font=font_body(size=10, bold=True),
                       alignment=center, border=border_all)
        elif coord == "I3":
            style_cell(cell, fill_color=GREEN,
                       font=font_body(size=10, bold=True),
                       alignment=right_wrap, border=border_all,
                       number_format='#,##0"원"')
        elif coord == "G4":
            style_cell(cell, fill_color=PALE_BLUE, font=font_body(size=10),
                       alignment=right_wrap, border=border_all,
                       number_format='#,##0"원"')
        elif coord == "F3":
            style_cell(cell, fill_color=LIGHT_GRAY,
                       font=font_body(size=9, italic=True),
                       alignment=center, border=border_all)
        elif coord == "F4":
            style_cell(cell, fill_color=LIGHT_GRAY,
                       font=font_body(size=9, italic=True),
                       alignment=center, border=border_all)

# 월별 데이터 블록 함수
def format_month_block(ws, title_row, data_rows, title_cols=1,
                       data_cols=range(2, 6), value_format='#,##0"원"',
                       highlight_last=False):
    """
    title_row: 헤더 행 번호 (A열은 제목, B~E열은 월 이름)
    data_rows: 데이터가 들어있는 행 리스트
    """
    # 타이틀 행
    style_cell(ws.cell(row=title_row, column=1),
               fill_color=NAVY, font=font_header(size=11),
               alignment=center, border=border_all)
    for col in data_cols:
        style_cell(ws.cell(row=title_row, column=col),
                   fill_color=LIGHT_BLUE, font=font_body(size=10, bold=True),
                   alignment=center, border=border_all)

    # 데이터 행
    for idx, r in enumerate(data_rows):
        is_last = (idx == len(data_rows) - 1) and highlight_last
        # A열: 항목명
        style_cell(ws.cell(row=r, column=1),
                   fill_color=YELLOW if is_last else LIGHT_BLUE,
                   font=font_body(size=10, bold=True),
                   alignment=left_wrap, border=border_all)
        # 데이터 열
        for col in data_cols:
            bg = YELLOW if is_last else (PALE_BLUE if r % 2 == 0 else WHITE)
            bold = is_last
            color = "C00000" if is_last else "000000"
            style_cell(ws.cell(row=r, column=col),
                       fill_color=bg,
                       font=font_body(size=10, bold=bold, color=color),
                       alignment=right_wrap, border=border_all,
                       number_format=value_format)


# 매출 블록 (A10:E13)
format_month_block(ws, 10, [11, 12, 13], highlight_last=True)

# 신규 매출 블록 (A15:E18)
format_month_block(ws, 15, [16, 17, 18], highlight_last=True)

# 재구매 매출 블록 (A20:E23)
format_month_block(ws, 20, [21, 22, 23], highlight_last=True)

# 재구매 건수 블록 (A25:E27)
format_month_block(ws, 25, [26, 27], value_format='#,##0')

# AOV 블록 (A29:E31)
format_month_block(ws, 29, [30, 31])

# 신규/재구매 고객 수 블록 (A35:E39)
format_month_block(ws, 35, [36, 37, 38, 39], value_format='#,##0')

# 재구매율 블록 (A40:E42)
format_month_block(ws, 40, [41, 42], value_format='0.0"%"')

# 재구매비 블록 (A44:E46)
format_month_block(ws, 44, [45, 46], value_format='0.00')

# 재구매 매출 비중 블록 (A52:E54)
format_month_block(ws, 52, [53, 54], value_format='0.0"%"')

# 재구매율 추이 블록 (A56:E58)
format_month_block(ws, 56, [57, 58], value_format='0.0"%"')

# 설명 영역 (A60:B83) - 노트/메모
for row_num in range(60, 84):
    for col in range(1, 3):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            is_title = row_num in [60, 61, 62, 66, 67, 68, 70, 78]
            style_cell(cell,
                       fill_color=LIGHT_GRAY if not is_title else LIGHT_BLUE,
                       font=font_body(size=10 if is_title else 9,
                                      bold=is_title, italic=not is_title),
                       alignment=left_wrap, border=border_all)

# 우측 설명 컬럼 (G2, G10:G33, G37, G57)
for row_num in [2, 10, 11, 12, 13, 14, 15, 22, 30, 32, 33, 37, 57]:
    cell = ws.cell(row=row_num, column=7)
    if cell.value is not None:
        style_cell(cell, fill_color=LIGHT_GRAY,
                   font=font_body(size=9, italic=True),
                   alignment=left_wrap, border=border_all)
    cell2 = ws.cell(row=row_num, column=8)
    if cell2.value is not None:
        style_cell(cell2, fill_color=LIGHT_GRAY,
                   font=font_body(size=9, italic=True),
                   alignment=left_wrap, border=border_all)

set_col_widths(ws, {
    "A": 28, "B": 14, "C": 14, "D": 14, "E": 14,
    "F": 18, "G": 32, "H": 22, "I": 20,
})
ws.freeze_panes = "A2"

# ========== 시트 5: 월별 신규구매타겟 ==========
ws = wb[sheets[4]]

style_cell(ws["I1"], fill_color=BLUE, font=font_header(size=10),
           alignment=center, border=border_all)

# 매출 블록 (A3:E6)
format_month_block(ws, 3, [4, 5, 6], highlight_last=True)
# 재구매 매출 (A8:E11)
format_month_block(ws, 8, [9, 10, 11], highlight_last=True)
# 신규구매자 매출 (A14:E17)
format_month_block(ws, 14, [15, 16, 17], highlight_last=True)
# 신규구매자 매출 2 (A20:E23)
format_month_block(ws, 20, [21, 22, 23], highlight_last=True)
# 신규구매자 매출전환율 (A25:E28)
format_month_block(ws, 25, [26, 27, 28], value_format='0.0"%"', highlight_last=True)

# 우측 비고 영역 (G8:G10)
for row_num in [8, 9, 10]:
    cell = ws.cell(row=row_num, column=7)
    if cell.value is not None:
        style_cell(cell, fill_color=LIGHT_GRAY,
                   font=font_body(size=9, italic=True),
                   alignment=left_wrap, border=border_all)

set_col_widths(ws, {
    "A": 28, "B": 14, "C": 14, "D": 14, "E": 14,
    "F": 14, "G": 32, "H": 22, "I": 18,
})
ws.freeze_panes = "A2"

# ========== 시트 6: 월별 목표달성 ==========
ws = wb[sheets[5]]

# 상단 연간 목표 (A3:D4)
for row_num in [3, 4]:
    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            if col == 1:
                style_cell(cell, fill_color=NAVY, font=font_header(size=11),
                           alignment=center, border=border_all)
            else:
                style_cell(cell, fill_color=GREEN,
                           font=font_body(size=11, bold=True, color="375623"),
                           alignment=center, border=border_all)

# 월 목표 테이블 헤더 (A8:H8)
apply_header(ws, "A8:H8")

# 월 목표 데이터 (A9:H15)
for row_num in range(9, 16):
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            if col == 1:  # 월 이름
                style_cell(cell, fill_color=LIGHT_BLUE,
                           font=font_body(size=10, bold=True),
                           alignment=center, border=border_all)
            elif col == 8:  # 비고
                style_cell(cell, fill_color=LIGHT_GRAY,
                           font=font_body(size=9, italic=True),
                           alignment=left_wrap, border=border_all)
            elif col == 7:  # 순수익
                style_cell(cell, fill_color=YELLOW,
                           font=font_body(size=11, bold=True, color="C00000"),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')
            else:
                bg = PALE_BLUE if row_num % 2 == 0 else WHITE
                style_cell(cell, fill_color=bg, font=font_body(size=10),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')

# A20 - 비고
style_cell(ws["A20"], fill_color=LIGHT_GRAY,
           font=font_body(size=9, italic=True),
           alignment=left_wrap)

set_col_widths(ws, {
    "A": 22, "B": 18, "C": 18, "D": 18, "E": 16,
    "F": 16, "G": 16, "H": 40,
})
set_row_heights(ws, {8: 32})
ws.freeze_panes = "A9"

# ========== 시트 7: 월별 매출 데이터 ==========
ws = wb[sheets[6]]

# 헤더 (B4:I4)
for col in range(2, 10):
    cell = ws.cell(row=4, column=col)
    if cell.value is not None:
        style_cell(cell, fill_color=NAVY, font=font_header(size=11),
                   alignment=center, border=border_all)

# A열 항목 + 좌측 테이블 (A5:D8)
for row_num in range(5, 9):
    for col in range(1, 5):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            if col == 1:
                style_cell(cell, fill_color=LIGHT_BLUE,
                           font=font_body(size=10, bold=True),
                           alignment=left_wrap, border=border_all)
            else:
                bg = YELLOW if row_num == 7 else (PALE_BLUE if row_num % 2 == 0 else WHITE)
                bold = row_num == 7
                color = "C00000" if row_num == 7 else "000000"
                style_cell(cell, fill_color=bg,
                           font=font_body(size=10, bold=bold, color=color),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')

# F열 항목 + 우측 테이블 (F5:I8)
for row_num in range(5, 9):
    for col in range(6, 10):
        cell = ws.cell(row=row_num, column=col)
        if cell.value is not None:
            if col == 6:
                style_cell(cell, fill_color=LIGHT_BLUE,
                           font=font_body(size=10, bold=True),
                           alignment=left_wrap, border=border_all)
            else:
                bg = YELLOW if row_num == 7 else (PALE_BLUE if row_num % 2 == 0 else WHITE)
                bold = row_num == 7
                color = "C00000" if row_num == 7 else "000000"
                style_cell(cell, fill_color=bg,
                           font=font_body(size=10, bold=bold, color=color),
                           alignment=right_wrap, border=border_all,
                           number_format='#,##0"원"')

set_col_widths(ws, {
    "A": 18, "B": 18, "C": 18, "D": 16, "E": 4,
    "F": 18, "G": 18, "H": 18, "I": 16,
})
set_row_heights(ws, {4: 28})
ws.freeze_panes = "A5"

# ========== 저장 ==========
wb.save(SRC)
print("\n[OK] saved")
