"""
더다 양식 엑셀 생성 모듈
- 원본 템플릿을 복사한 뒤 데이터만 채워넣어서 서식/함수 보존
"""

import os
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from dotenv import load_dotenv

# .env 먼저 로드 (서버 배포 시 경로 오버라이드 위해)
_ENV_PATH = Path(__file__).parent / ".env"
load_dotenv(_ENV_PATH, override=True)

# 경로 설정 - 환경변수 DADA_FOLDER/DADA_TEMPLATE로 오버라이드 가능 (서버 배포용)
_DEFAULT_DADA_FOLDER = Path("C:/Users/osh80/OneDrive/바탕 화면/사업/더다 3pl/더다 양식")
_DEFAULT_TEMPLATE = _DEFAULT_DADA_FOLDER / "더다냉동물류 발주양식 25.7.8.xlsx"

DADA_FOLDER = Path(os.getenv("DADA_FOLDER") or _DEFAULT_DADA_FOLDER)
TEMPLATE_FILE = Path(os.getenv("DADA_TEMPLATE") or _DEFAULT_TEMPLATE)

# 더다 양식 컬럼 순서 (A~K)
DADA_COLUMNS = [
    "고객주문번호",       # A
    "받는분성명",          # B
    "받는분전화번호",       # C
    "받는분기타연락처",     # D
    "받는분우편번호",       # E
    "받는분주소(전체,분할)", # F
    "상품명",             # G
    "상품상세",            # H
    "내품수량",            # I
    "배송메세지1",         # J
    "송장번호",            # K
]


def get_today_filename():
    """오늘 날짜로 파일명 생성 (예: 더다냉동물류 발주양식 26.4.17.xlsx)"""
    now = datetime.now()
    date_str = f"{now.year % 100}.{now.month}.{now.day}"
    return f"더다냉동물류 발주양식 {date_str}.xlsx"


def create_dada_file(df, output_path=None):
    """
    원본 템플릿을 복제한 뒤 데이터만 채워넣기
    - 서식, 열너비, 테두리, 숨김컬럼 등 전부 보존됨

    Args:
        df: 더다 양식 컬럼을 가진 DataFrame
        output_path: 저장할 파일 경로 (None이면 오늘 날짜로 자동 생성)

    Returns:
        생성된 파일 경로 (Path)
    """
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"템플릿 파일을 찾을 수 없음: {TEMPLATE_FILE}")

    if output_path is None:
        output_path = DADA_FOLDER / get_today_filename()
    output_path = Path(output_path)

    # 같은 이름 파일이 이미 있으면 백업 (덮어쓰기 방지)
    if output_path.exists():
        now = datetime.now()
        backup_name = f"{output_path.stem}.backup_{now:%H%M%S}{output_path.suffix}"
        backup_path = output_path.parent / backup_name
        shutil.copy2(output_path, backup_path)

    # 1) 원본 템플릿을 그대로 복사
    shutil.copy2(TEMPLATE_FILE, output_path)

    # 2) 복사된 파일 열기
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # 3) 기존 데이터 행 지우기 (2행부터, 11개 컬럼만)
    #    ※ 12번 이후 컬럼은 원본에 있던 그대로 보존
    if ws.max_row >= 2:
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, len(DADA_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None

    # 4) 새 데이터 쓰기 (2행부터)
    df = df.fillna("")
    for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(DADA_COLUMNS, start=1):
            if col_name not in df.columns:
                continue
            value = row[col_name]
            # 빈 문자열은 아예 비워두기
            if value == "" or value is None:
                continue
            ws.cell(row=row_offset, column=col_idx, value=value)

    wb.save(output_path)
    return output_path


def validate_dada_file(filepath, expected_count=None):
    """
    생성된 더다 양식 파일 검증
    - 파일 존재 확인
    - 헤더 일치 확인
    - 필수 컬럼에 값 있는지 확인
    - 행 수가 expected_count와 일치하는지

    Returns:
        (is_valid: bool, issues: list[str])
    """
    filepath = Path(filepath)
    issues = []

    if not filepath.exists():
        return False, [f"파일이 존재하지 않음: {filepath}"]

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        return False, [f"파일을 열 수 없음: {e}"]

    ws = wb.active

    # 헤더 확인
    headers = [ws.cell(row=1, column=i).value for i in range(1, len(DADA_COLUMNS) + 1)]
    for i, (expected, actual) in enumerate(zip(DADA_COLUMNS, headers), 1):
        if expected != actual:
            issues.append(f"헤더 불일치 (열 {i}): 예상={expected!r}, 실제={actual!r}")

    # 데이터 행 수 확인
    data_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        # A열(주문번호)에 값이 있으면 데이터 행으로 카운트
        if ws.cell(row=row_idx, column=1).value:
            data_rows += 1

    if expected_count is not None and data_rows != expected_count:
        issues.append(f"행 수 불일치: 예상={expected_count}, 실제={data_rows}")

    # 필수 컬럼 비어있는지 확인
    required_columns = {
        1: "고객주문번호",
        2: "받는분성명",
        3: "받는분전화번호",
        6: "받는분주소",
        7: "상품명",
        9: "내품수량",
    }

    for row_idx in range(2, ws.max_row + 1):
        if not ws.cell(row=row_idx, column=1).value:
            continue  # 빈 행 스킵
        for col_idx, col_name in required_columns.items():
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None or value == "":
                issues.append(f"행 {row_idx}: {col_name} 비어있음")

    is_valid = len(issues) == 0
    return is_valid, issues


if __name__ == "__main__":
    # 테스트: 샘플 데이터 만들어서 엑셀 생성
    sample = pd.DataFrame([
        {
            "고객주문번호": "TEST-001",
            "받는분성명": "홍길동",
            "받는분전화번호": "010-1234-5678",
            "받는분기타연락처": "",
            "받는분우편번호": "12345",
            "받는분주소(전체,분할)": "서울시 강남구 테헤란로 123",
            "상품명": "불닭 닭가슴살 현미밥",
            "상품상세": "",
            "내품수량": 3,
            "배송메세지1": "부재시 경비실",
            "송장번호": "",
        }
    ])

    output = create_dada_file(sample, DADA_FOLDER / "테스트_더다양식.xlsx")
    print(f"생성됨: {output}")

    is_valid, issues = validate_dada_file(output, expected_count=1)
    print(f"검증 결과: {'통과' if is_valid else '실패'}")
    for issue in issues:
        print(f"  - {issue}")
