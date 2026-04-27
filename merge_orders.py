"""
헤비로버 주문 자동 취합 스크립트
- 스마트스토어 + 카페24 주문 → 더다 3PL 발주양식 엑셀 생성
"""

import io
import os
import sys
from datetime import datetime

import msoffcrypto
import openpyxl
import pandas as pd


def read_smartstore(filepath, password="1234"):
    """스마트스토어 암호화 엑셀 읽기"""
    with open(filepath, "rb") as f:
        ms = msoffcrypto.OfficeFile(f)
        ms.load_key(password=password)
        decrypted = io.BytesIO()
        ms.decrypt(decrypted)

    decrypted.seek(0)
    df = pd.read_excel(decrypted, engine="openpyxl", header=1)

    # 더다 양식에 맞게 컬럼 매핑
    result = pd.DataFrame()
    result["고객주문번호"] = df["상품주문번호"].astype(str)
    result["받는분성명"] = df["수취인명"]
    result["받는분전화번호"] = df["수취인연락처1"]
    result["받는분기타연락처"] = df.get("수취인연락처2", "")
    result["받는분우편번호"] = df.get("우편번호", "")
    result["받는분주소(전체,분할)"] = df["통합배송지"]
    result["상품명"] = df["옵션정보"]
    result["상품상세"] = ""
    result["내품수량"] = df["수량"]
    result["배송메세지1"] = df["배송메세지"]
    result["송장번호"] = ""
    result["출처"] = "스마트스토어"

    return result


def read_cafe24(filepath):
    """카페24 CSV 읽기"""
    # 여러 인코딩 시도
    for enc in ["utf-8", "utf-8-sig", "euc-kr", "cp949"]:
        try:
            df = pd.read_csv(filepath, encoding=enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    # 더다 양식에 맞게 컬럼 매핑
    result = pd.DataFrame()
    result["고객주문번호"] = df["주문번호"].astype(str)
    result["받는분성명"] = df["수령인"]
    result["받는분전화번호"] = df["핸드폰"]
    result["받는분기타연락처"] = df.get("수령지전화", "")
    result["받는분우편번호"] = df.get("우편번호", "")
    result["받는분주소(전체,분할)"] = df["주소"]
    result["상품명"] = df["옵션"]
    result["상품상세"] = ""
    result["내품수량"] = df["수량"]
    result["배송메세지1"] = df.get("비고", "")
    result["송장번호"] = ""
    result["출처"] = "카페24"

    return result


def create_dada_excel(smartstore_df, cafe24_df, output_path):
    """더다 3PL 발주양식 엑셀 생성"""
    # 두 플랫폼 데이터 합치기
    combined = pd.concat([smartstore_df, cafe24_df], ignore_index=True)

    # NaN을 빈 문자열로
    combined = combined.fillna("")

    # 더다 양식 컬럼 순서 (출처 컬럼 제외)
    dada_columns = [
        "고객주문번호",
        "받는분성명",
        "받는분전화번호",
        "받는분기타연락처",
        "받는분우편번호",
        "받는분주소(전체,분할)",
        "상품명",
        "상품상세",
        "내품수량",
        "배송메세지1",
        "송장번호",
    ]

    output_df = combined[dada_columns]

    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더 쓰기
    for col_idx, col_name in enumerate(dada_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 데이터 쓰기
    for row_idx, row in output_df.iterrows():
        for col_idx, col_name in enumerate(dada_columns, 1):
            value = row[col_name]
            ws.cell(row=row_idx + 2, column=col_idx, value=value)

    wb.save(output_path)
    return len(output_df)


def main():
    # 오늘 날짜 (예: 26.4.17)
    now = datetime.now()
    today = f"{now.year % 100}.{now.month}.{now.day}"

    # 파일 경로 설정
    desktop = os.path.expanduser("~/OneDrive/바탕 화면")

    # 스마트스토어 파일 찾기 (가장 최신 파일)
    smartstore_files = [
        f for f in os.listdir(desktop)
        if f.startswith("스마트스토어_전체주문발주발송관리") and f.endswith(".xlsx")
    ]
    smartstore_files.sort(reverse=True)

    # 카페24 파일 찾기
    cafe24_files = [
        f for f in os.listdir(desktop)
        if f.endswith("_orders.csv")
    ]
    cafe24_files.sort(reverse=True)

    if not smartstore_files:
        print("스마트스토어 주문 파일을 찾을 수 없습니다.")
        return
    if not cafe24_files:
        print("카페24 주문 파일을 찾을 수 없습니다.")
        return

    smartstore_path = os.path.join(desktop, smartstore_files[0])
    cafe24_path = os.path.join(desktop, cafe24_files[0])

    print(f"스마트스토어 파일: {smartstore_files[0]}")
    print(f"카페24 파일: {cafe24_files[0]}")

    # 데이터 읽기
    ss_df = read_smartstore(smartstore_path)
    c24_df = read_cafe24(cafe24_path)

    print(f"스마트스토어 주문: {len(ss_df)}건")
    print(f"카페24 주문: {len(c24_df)}건")

    # 더다 양식 엑셀 생성
    output_dir = os.path.join(desktop, "사업", "더다 3pl", "더다 양식")
    output_filename = f"더다냉동물류 발주양식 {today}.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    total = create_dada_excel(ss_df, c24_df, output_path)

    print(f"\n더다 발주양식 생성 완료!")
    print(f"파일: {output_path}")
    print(f"총 {total}건 (스마트스토어 {len(ss_df)} + 카페24 {len(c24_df)})")


if __name__ == "__main__":
    main()
